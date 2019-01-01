#!/usr/bin/env python3

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

course = load_workbook('courses.xlsx')

def combine():

    #get table
    student = course.get_sheet_by_name('students')
    time = course.get_sheet_by_name('time')

    #create table combine
    combine = course.create_sheet('combine')
    combine.append(['创建时间', '课程名称', '学习人数', '学习时间'])

    #combine
    for stu in student.values:
        if stu[1] != '课程名称':
            for t1 in time.values:
                if stu[1] == t1[1]:
                    combine.append(list(stu) + [t1[2]])

    course.save('courses.xlsx')

def split():
    yearlist = []
    combine = course.get_sheet_by_name('combine')
    for com in combine.values:
        if com[1] != '课程名称':
            yearlist.append(com[0].strftime('%Y'))
    years = set(yearlist)

    for yer in years:
        wb_tmp = Workbook()
        wb_tmp.remove(wb_tmp.active)
        ws = wb_tmp.create_sheet(yer)

        ws.append(['创建时间', '课程名称', '学习人数', '学习时间'])
        for item in combine.values:
            if item[1] != '课程名称':
                if yer == item[0].strftime('%Y'):
                    ws.append(item)
        # for val in ws.values:
        #     print(val[1], end='')
        #
        # print()
        print(wb_tmp)
        wb_tmp.save('{}.xlsx'.format(yer))

if __name__ == '__main__':
    combine()
    split()
